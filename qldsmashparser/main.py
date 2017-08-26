"""
Copyright (c) 2017 Megan Leet (Zowayix)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

You'll need to install BeautifulSoup 4, dateutil, and openpyxl (and also Pillow
if you want draw_table.py to work)
Please ignore any pylint warnings about a space being after a bracket
where there's something like (**args), that's because NetBeans keeps fucking
it up when I press autoformat
And yes I use tabs for indentation, shove the PEP-8 guide up your arse (I regret writing
this in Python but it's too late now)
I'll write more doco later hey
This shouldn't be called "main" or should it
Task list:
	- Hmm... what if a chart of how set win % or other
	get_player_aggregate_data stuff changes over time, e.g. per day or per event?
	- Calculate "Ghost numbers" and stuff like that
		- Need to think of an algorithm
	- Better error handling when I try to get a player or tournament page that
	doesn't exist
	- Refactor things to be a bit cleaner and objecty?
		- There's also a lot of duplicated code which could be made into functions or whatevs
		- Move some logic around from OpponentMatchupData.create into public methods
	- Rate limits (until then, please use this responsibly)
	- Is caching the whole HTML page overkill? (tournament_cache ungzipped is 117MB atm)
	- Make variable names more consistent, I guess
	- All TODOs in the comments and shit
	- Make sure draw_table.py still works with the new version of Pillow
	- Invalidate player cache every week perhaps
	- Should be nice and cache get_all_players
	- Put the output to spreadsheet thing into a new module
"""

import cProfile
from core import *
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import PatternFill

#People from ACT who show up to stuff
ZOWAYIX20_CANDIDATES = [('ACT', x) for x in [
	'Waveguider', 'Sriks', 'Pazx', 'Maplemage', 'Glacier', 'Calamity', 'Skylar', 'Joey',
	'Cherry', 'SJ AAAce3', 'Tila', 'Zowayix', 'Eight', 'Flyingcat', 'Zac', 'GVol3',
	'Frog', 'Zeus', 'FleetCMDR', 'BreakVG', 'Rose', 'Gracetail', 'quen', 'TJSlash'
	]]
		
#The NSW people wanted some stats too.
#Going by the MM40 as of Mar 21 2017 (last updated Feb 4 2017)							
MM_40 = [('NSW', x) for x in ['Luco', 'Killy', 'Jeese', 'MM', 'Shaya', 'Joe', 'Kristoph', 
	'Boundaries', 'Dr. Ainuss', 'w00tkins', 'Invisi', 'Buster', 'Enn', 'Inusaki', 'Emansaur',
	'Atyeo', 'BJSchoey', 'Scarpian', 'Trojans', '4par', 'Thorpenado', 'Fruit',
	'Hans', 'Zedi', 'Lanatra', 'NINjA', 'Lumi', 'Ash', 'Zesco', 'SpacemanBad',
	'Kelp', 'Thrillhouse', 'Benjam', 'Sylem', 'Quy', 'Revan', 'Struz', 'Naomi Campbelltown',
	'Bjay', 'Zephyr']]

#TODO This is the olD PR lol
AUS_PR = [('SA', 'Ghost'), ('VIC', 'Extra'), ('ACT', 'Waveguider'), ('NSW', 'Luco'),
('QLD', 'Jezmo'), ('QLD', 'Jaice'), ('VIC', 'Earl'), ('WA', 'Poppt1'), ('NSW', 'MM'),
('VIC', 'Revax'), ('VIC', 'Ignis'), ('NSW', 'SaucyDancer'), ('VIC', 'Duon'), ('VIC', 'Boozer')]


def get_data_against_player_list(player, player_list, date_limit=None):
	"""
	For a given player, gets some information about how they go against
	a list of other players (tuples with region, name)
	"""
	data = {}
	for opponent in player_list:
		data[opponent] = player.get_opponent_data('SSBU', Player(*opponent), date_limit)
	return data

def get_player_matchup_table(prop, player_list, date_limit=None):
	"""
	For a given list of players (tuples of region, name), returns a dictionary
	with a player as each key and a dict with each opponent's matchup data against
	that player as value
	"""
	data = {}
	for player in player_list:
		opponent_data = get_data_against_player_list(Player(*player), player_list, date_limit)
		data[player] = {key: getattr(value, prop) for key, value in opponent_data.items()}
	return data


def write_table_to_worksheet(table, worksheet, number_format='0.##', 
							 colour_scale=(0, 100), aggregate=None):
	"""
	I'll explain the parameters later, can't be buggered atm tbh fam
	Yeah okay I guess this is a bit too long in that it does have
	too many local variables
	"""
	column_numbers = {}
	
	row_num = 1
	
	for row, columns in table.items():		
		row_num += 1
		worksheet.cell(row=row_num, column=1).value = row
		
		for column, value in columns.items():
			if column in column_numbers:
				col_num = column_numbers[column]
			else:
				col_num = max(column_numbers.values()) + 1 if column_numbers else 2
				header = worksheet.cell(row=1, column=col_num)
				header.value = column
				#I actually want 270, which is perfectly valid in a
				#spreadsheet, but openpyxl doesn't think so, and maybe
				#I should submit them a bug report for that
				#Also I can't autosize rows so I might as well not
				#bother doing it in Python
				#header.alignment = Alignment(text_rotation=90)
				column_numbers[column] = col_num
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = value
			cell.number_format = number_format
			
	end_row = row_num
	end_col = max(column_numbers.values()) if column_numbers else 1
	end_col_letter = worksheet.cell(column=end_col, row=1).column
	
	if aggregate is not None and end_row > 1 and end_col > 1:
		for i in range(2, end_row + 1):
			aggregate_cell = worksheet.cell(row=i, column=end_col + 1)
			aggregate_cell.value = '={0}(B{1}:{2}{1}'.format(aggregate, i, end_col_letter)
			aggregate_cell.number_format = number_format

	if colour_scale is not None and end_row > 1 and end_col > 1:
		rule = ColorScaleRule(
							  start_type='num', end_type='num',
							  start_value=colour_scale[0], end_value=colour_scale[1],
							  start_color='FF0000', end_color='00FF00')
		end_cell = worksheet.cell(row=end_row, column=end_col)
		address = 'B2:' + end_cell.coordinate
		worksheet.conditional_formatting.add(address, rule)
		
		gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
		na_rule = CellIsRule(operator='=', formula=['"N/A"'], fill=gray_fill)
		worksheet.conditional_formatting.add(address, na_rule)
		
def write_player_matchup_data_to_workbook(date_limit, player_list, props):
	"""
	Returns a new openpyxl Workbook object (you can save it yourself) with
	as many sheets containing OpponentMatchupData information for all players
	in player_list against each other, depending on
	what you put in props. Props is a list of tuples with (property name of
	OpponentMatchupData object, tuple for the limits of data for conditional
	formatting (e.g. (0, 100) to make a colour scale where 0 is one colour
	and 100 is the other) or None to not use conditional formatting, name of
	an Excel function to put on the end of each row eg AVERAGE or SUM)
	"""
	book = Workbook()
	book.remove(book.active)
	for prop in props:
		sheet = book.create_sheet(prop[0])
		print('Writing %s' % prop[0])
		write_table_to_worksheet(get_player_matchup_table(prop[0], player_list, date_limit),
								 sheet, colour_scale=prop[1], aggregate=prop[2])
	return book
	#Things you still have to do yourself because openpyxl won't: Autosize columns,
	#autosize rows for that matter, rotate headers if wanted, sort
	

def get_player_aggregate_data(player, game, date_limit=None):
	"""
	Returns a dict containing a summary of sorts for this player for a certain game, optionally
	over a certain time period, e.g. set win rate %, total sets won, total sets played, and so
	on and such for individual games as well
	"""
	def __set_filter(_set):
		if _set.game != game:
			return False
		if _set.forfeit:
			return False
		if date_limit is not None:
			return _set.date > date_limit
		return True
	
	#While the rest of the code looks like you don't have to call list() on this
	#line and just leave sets as being a generator, this results in everyone having
	#a 0% win rate and 0 total games played, and I can't be stuffed figuring
	#out why that is so I'm gonna say because Python
	sets = list(filter(__set_filter, player.get_sets()))
	total_set_count = len(list(sets))
	if total_set_count == 0:
		return {x: 0 for x in ['Set wins', 'Total sets played', 'Set win %', 'Game wins',
			'Total games played', 'Game win%']}
		
	set_win_count = len([the_set for the_set in sets if the_set.won_by_player(player)])
	set_win_rate = (set_win_count / total_set_count) * 100
	
	game_win_count = sum(the_set.winner.games_won for the_set in sets
						 if the_set.won_by_player(player))
	game_lose_count = sum(the_set.winner.games_won for the_set in sets
						  if not the_set.won_by_player(player))
	total_game_count = game_win_count + game_lose_count
	game_win_rate = (game_win_count / total_game_count) * 100
	
	return {'Set wins': set_win_count, 'Total sets played': total_set_count, 'Set win %': set_win_rate,
		'Game wins': game_win_count, 'Total games played': total_game_count, 'Game win %': game_win_rate}

def get_aggregate_data_for_player_list(player_list, game, date_limit=None):
	"""Returns a dict containing each player's result of get_player_aggregate_data"""
	return {player[1]: get_player_aggregate_data(Player(*player), game, date_limit)
		for player in player_list}

def get_problem_characters_for_player(player, game):
	"""
	Shows which characters player most often loses to in a game.
	This isn't 100% accurate because character data doesn't go down to the
	level of each individual game, and also isn't always filled in"""
	tally = {}
	sets = [_set for _set in player.get_sets() if _set.game == game]
	for _set in sets:
		won = _set.won_by_player(player)
		opponent_chars = _set.loser.characters if won else _set.winner.characters
		for char in opponent_chars:
			if char not in tally:
				tally[char] = {'wins': 0, 'losses': 0}
			if won:
				tally[char]['wins'] += 1
			else:
				tally[char]['losses'] += 1
	
	return {key: (value['losses'] / (value['wins'] + value['losses']) * 100)
		for key, value in tally.items()}

def get_upsets_by_elo(player):
	"""
	Returns a list of SmashSet objects for a certain player where that player
	won and it was considered an "upset", considering a player with more Elo for
	that game to be the normally better player
	"""
	sets = [_set for _set in player.get_sets() if _set.won_by_player(player)]
	def __is_upset(_set):
		opponent_elo = _set.loser.get_elo(_set.game)
		if opponent_elo is None:
			return False
		return player.get_elo(_set.game) < opponent_elo
	return list(filter(__is_upset, sets))

def get_upsets_by_overall_win_rate(player):
	"""
	Returns a list of SmashSet objects for a certain player where that player
	won and it was considered an "upset", considering a player with a higher ratio
	of sets won to total sets played over their career in that game to be the
	better player
	"""
	win_rates = {}
	
	sets = [_set for _set in player.get_sets() if _set.won_by_player(player)]
	def __is_upset(_set):
		if _set.game not in win_rates:
			win_rates[_set.game] = get_player_aggregate_data(player, _set.game)['Set win %']
		
		if _set.loser.region is None:
			#They're not in the database, assume that they're nobody important and
			#therefore this isn't an upset
			return False
		opponent_win_rate = get_player_aggregate_data(_set.loser, _set.game)['Set win %']
		return win_rates[_set.game] < opponent_win_rate
	#return list(filter(__is_upset, sets))
	return [_set for _set in sets if __is_upset(_set)]

def list_upsets_by_elo(player):
	"""
	Prints the results of get_upsets_by_elo nicely.
	"""
	print('Upsets for {0} from {1} ({2})'.format(player.name, player.region,
		  ', '.join('{0}: {1}'.format(game, player.get_elo(game)) for
		  game in player.get_games_played())))
	for upset in get_upsets_by_elo(player):
		print('{0} {1} - {2} vs. {3} ({4})'.format(
			  upset.tournament.name, upset.bracket_name, upset.get_formatted_round(True),
			  upset.loser.name, upset.loser.get_elo(upset.game)))

def list_upsets_by_overall_win_rate(player):
	"""
	Prints the results of get_upsets_by_overall_win_rate nicely. TODO I was supposed to genericize
	this so that the same logic is used to print get_upsets_by_elo as well WHOOPS
	"""
	print('Upsets for {0} from {1} ({2})'.format(player.name, player.region,
		  ', '.join('{0}: {1:.3g}%'.format(game, get_player_aggregate_data(player, game)['Set win %']) for
		  game in player.get_games_played())))
	for upset in get_upsets_by_overall_win_rate(player):
		print('{0} {1} - {2}\n\tUsed {3} vs. {4} ({5:.3g}%) {6}'.format(
			  upset.tournament.name, upset.bracket_name, upset.get_formatted_round(True),
			  ', '.join(upset.winner.characters), upset.loser.name,
			  get_player_aggregate_data(upset.loser, upset.game)['Set win %'],
			  ('using ' + ', '.join(upset.loser.characters)) if upset.loser.characters else ''))
			  
def list_tournament_performances(player):
	brackets = {}
	for _set in player.get_sets():
		if _set.bracket in brackets:
			brackets[_set.bracket].append(_set)
		else:
			brackets[_set.bracket] = [_set]
	
	for bracket, sets in brackets.items():
		print('{0} - {1}\t\t: {2} - {3}'.format(
			  bracket.tournament.name, bracket.name,
			  len(list(_set for _set in sets if _set.won_by_player(player))), 
			  len(list(_set for _set in sets if not _set.won_by_player(player))), ))
			  
def write_aggregate_data_to_workbook(player_list, game, date_limit, filename):
	data = get_aggregate_data_for_player_list(player_list, game, date_limit)
	wb = Workbook()
	ws = wb.active
	write_table_to_worksheet(data, ws, colour_scale=None)
	wb.save(filename)
	
def get_tournament_attendance(player_list, game, start_date, end_date, minimum=1):
	def __get_tourney_count(player):
		sets = [_set for _set in player.get_sets() if _set.date >= start_date and _set.date <= end_date]
		tournaments = {_set.tournament for _set in sets if _set.game == game}
		return len(tournaments)
	
	return {player: __get_tourney_count(Player(*player) if isinstance(player, tuple) else player) 
		for player in player_list}

def get_all_players(region):
	def __get_players_from_byregion_page(page):
		list_element = page.find(class_='table-hover')
		if list_element is None:
			return None
		return [Player.from_url(p['data-link']) for p in list_element('tr')]
	first_page_url = 'https://qldsmash.com/Players/ByRegion/{0}/1'.format(quote(region))
	first_page = BeautifulSoup(get_http(first_page_url), 'lxml')
	player_list = __get_players_from_byregion_page(first_page)

	next_page_button = first_page.find(class_='PagedList-skipToNext')
	while next_page_button is not None:
		#I don't like joining URLs manually like that, but eh, it'll work
		next_page_url = 'https://qldsmash.com/' + next_page_button.find('a')['href']
		next_page = BeautifulSoup(get_http(next_page_url), 'lxml')
		player_list += __get_players_from_byregion_page(next_page)
		next_page_button = next_page.find(class_='PagedList-skipToNext')

	return player_list

#Ignore all the commented out crap below. This is basically just because I don't have a CLI
#or GUI to run stuff from, so because I'm lazy and sloppy I just enter something I want to
#run here and call it and yeah

#print(max(s.date for s in Player('ACT', 'Zowayix').get_sets()))

#for k, v in sorted({player.name: player.get_elo('SSBU') for player in get_all_players('ACT')
#	if 'SSBU' in player.get_games_played()}.items(), key=lambda p: p[1], reverse=True):
#	print(k, v)

#attendance = get_tournament_attendance(get_all_players('ACT'), 'SSBU', date(2017, 3, 23), date(2017, 7, 21))
#for key, value in sorted(attendance.items(), key=lambda item: item[1], reverse=True):
#	if value > 0:
#		print(key.name, value)
	
#cProfile.run('list_tournament_performances(Player("ACT", "Zowayix"))')
#print(Tournament('ACT', 'Capital Smash 22')['Smash 4 poolz'].get_sets())
			  
#for key, value in get_aggregate_data_for_player_list(AUS_PR, 'SSBU', date(2017, 1, 1)).items():
#	print('{0}: {1}'.format(key, value))

#list_upsets_by_elo(Player('ACT', 'Zowayix'))
#cProfile.run("list_upsets_by_overall_win_rate(Player('ACT', 'Zowayix'))")

#blah = get_problem_characters_for_player(Player('ACT', 'Zowayix'), 'SSBU')
#print('\n'.join(["{0}\t\t: {1}".format(*item) for item in sorted(blah.items(),
#	  key=lambda item: item[1], reverse=True)]))

#print(get_player_aggregate_data(Player('ACT', 'Zowayix'), date(2016, 12, 7)))
#data = get_aggregate_data_for_player_list(AUS_PR, 'SSBU', date(2016, 12, 7))
#cProfile.run("write_aggregate_data_to_workbook(MM_40, 'SSBU', date(2017, 1, 1), 'aggregated_MM40_data.xlsx')")

#write_table_to_json(get_act_matchup_table('set_win_rate', date(2016, 12, 7)), 'table.json')
#print(get_data_against_player_list(Player('ACT', 'Zowayix'), 
#	RELEVANT_ACT_PLAYERS, date(2016, 12, 7)))
